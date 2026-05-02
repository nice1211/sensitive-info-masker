"""
Excel 文件处理器 (.xlsx)
遍历所有工作表的单元格进行脱敏/还原
"""
from openpyxl import load_workbook
from core.detector import SensitiveDetector
from core.masker import Masker


def mask_xlsx_file(filepath, detector: SensitiveDetector, items=None):
    """
    对 xlsx 文件执行脱敏
    items: 可选，前端传来的已筛选检测项列表。为 None 时自动检测。
    返回: (Workbook 对象, mapping)
    """
    wb = load_workbook(filepath)
    masker = Masker()

    # 构建允许脱敏的 value 集合
    allowed_values = {item['value'] for item in items} if items is not None else None

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    detections = detector.detect(cell.value)
                    if allowed_values is not None:
                        detections = [d for d in detections if d['value'] in allowed_values]
                    if detections:
                        masked, _ = masker.mask_text(cell.value, detections)
                        cell.value = masked

    return wb, masker.get_mapping()


def unmask_xlsx_file(filepath, mapping):
    """
    对 xlsx 文件执行还原
    返回: Workbook 对象
    """
    wb = load_workbook(filepath)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell.value = Masker.unmask_text(cell.value, mapping)

    return wb
